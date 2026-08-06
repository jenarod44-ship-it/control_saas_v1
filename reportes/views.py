from django.contrib import messages
from django.shortcuts import redirect
from core.utils.asistencia import calcular_estado_asistencia

from core.services.prenomina_service import ResumenPrenomina

from core.excel.encabezado import (
    escribir_encabezado_reporte,
    escribir_encabezados,
)

from reportes.services.permisos import (
    obtener_resultados_permisos,
)

from reportes.adapters.permisos_excel import (
    construir_reporte_permisos,
)

from core.excel.estilos import (
    ALINEACION_CENTRO,
    ALINEACION_DERECHA,
    ALINEACION_IZQUIERDA,
    BORDE_FINO,
    FUENTE_NEGRITA,
    RELLENO_ALERTA,
    RELLENO_ERROR,
    RELLENO_GRIS,
    RELLENO_INFORMATIVO,
    RELLENO_OK,
)

from core.excel.impresion import configurar_impresion
from core.excel.respuesta import crear_respuesta_excel
from core.excel.encabezado import (
    escribir_encabezado_reporte,
    escribir_encabezados,
)
from core.excel.reporte import crear_reporte_excel
from reportes.adapters.asistencia_excel import (
    construir_reporte_asistencia,
)



from core.excel.impresion import configurar_impresion
from core.excel.respuesta import crear_respuesta_excel
from django.shortcuts import render
from asistencia.models import Asistencia
from nucleo.models import Empleado
from core.services.asistencia_service import calcular_estado_asistencia
from core.services.incidencias import generar_incidencias_por_rango
from core.models import Incidencia
from asistencia.models import Movimiento
from nucleo.models import Empleado
import csv
from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Prefetch
from core.decorators import solo_operativo
from core.utils.asistencia import debe_generar_falta
from core.utils.laboral import es_dia_laboral
from core.utils.asistencia import debe_generar_falta, es_tiempo_extra
from datetime import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from openpyxl.styles import Alignment
from core.models import IncidenciaDia
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment

from core.excel.encabezado import escribir_encabezado_reporte
from core.excel.estilos import (
    ALINEACION_CENTRO,
    ALINEACION_ENVUELTA,
    ALINEACION_IZQUIERDA,
    BORDE_FINO,
    FUENTE_ENCABEZADO,
    FUENTE_NEGRITA,
    RELLENO_ALERTA,
    RELLENO_ERROR,
    RELLENO_GRIS,
    RELLENO_INFORMATIVO,
    RELLENO_OK,
    RELLENO_TITULO,
)
from core.excel.impresion import configurar_impresion
from core.excel.respuesta import crear_respuesta_excel
from core.services.prenomina_service import ResumenPrenomina
from nucleo.models import Departamento, Empleado
from core.services.asistencia_service import obtener_tiempo_extra


def calcular_estado_asistencia(empleado, fecha):

    print("DEPTO:", empleado.departamento)
    print("TRABAJA FINES:", getattr(empleado.departamento, "trabaja_fines_semana", "NO EXISTE"))

    asistencia = empleado.asistencia_set.filter(fecha=fecha).first()

    dia = fecha.weekday()

    trabaja_fines = (
        empleado.departamento and
        getattr(empleado.departamento, "trabaja_fines_semana", False)
    )

    # 🔥 obtener asistencia
    asistencia = empleado.asistencia_set.filter(fecha=fecha).first()

    # 🔥 SI ES FIN DE SEMANA
    if dia in [5, 6]:

        # 🟢 SEGURIDAD → flujo normal
        if trabaja_fines:
            pass

        # 🔵 OTROS EMPLEADOS
        else:
            if asistencia:
                return "TIEMPO_EXTRA"

            return "NO_LABORAL"
    
    if not asistencia:
        return "FALTA"

    entrada = asistencia.hora_entrada
    salida = asistencia.hora_salida

    if es_tiempo_extra(empleado, fecha):
        return "TIEMPO_EXTRA"

    hora_limite_retardo = time(8, 15)

    if entrada:

        if entrada <= hora_limite_retardo:
            estado = "OK"
        else:
            estado = "RETARDO"

        if not salida:
            estado = "INCOMPLETO"

        return estado

    return "FALTA"

from datetime import datetime

def calcular_tiempo(salida, regreso):

    if not salida or not regreso:
        return ""

    t_salida = datetime.combine(datetime.today(), salida)
    t_regreso = datetime.combine(datetime.today(), regreso)

    diferencia = t_regreso - t_salida

    horas = diferencia.seconds // 3600
    minutos = (diferencia.seconds % 3600) // 60

    return f"{horas:02d}:{minutos:02d}"

from core.utils.asistencia import debe_generar_falta, es_tiempo_extra
from datetime import time


def calcular_estado_asistencia(empleado, fecha):

    incidencia_dia = IncidenciaDia.objects.filter(
        empleado=empleado,
        fecha=fecha,
    ).first()

    if incidencia_dia:
        return incidencia_dia.tipo

    asistencia = empleado.asistencia_set.filter(
        fecha=fecha,
    ).first()

    # Empleados sin control de horario:
    # no se evalúan días laborales ni retardos.
    if not empleado.control_horario:

        if not asistencia or not asistencia.hora_entrada:
            return "FALTA"

        if not asistencia.hora_salida:
            return "INCOMPLETO"

        return "OK"

    # Solo empleados con control de horario
    # se clasifican como día no laboral.
    if not debe_generar_falta(empleado, fecha):
        return "NO_LABORAL"

    if not asistencia:
        return "FALTA"

    entrada = asistencia.hora_entrada
    salida = asistencia.hora_salida

    if es_tiempo_extra(empleado, fecha):
        return "TIEMPO_EXTRA"

    hora_limite_retardo = time(8, 15)

    if entrada:

        if entrada <= hora_limite_retardo:
            estado = "OK"
        else:
            estado = "RETARDO"

        if not salida:
            estado = "INCOMPLETO"

        return estado

    return "FALTA"

def calcular_incidencias_asistencia(empleado, fecha):

    incidencias = []

    incidencias_qs = Incidencia.objects.filter(
        empleado=empleado,
        fecha_inicio__lte=fecha,
        fecha_fin__gte=fecha
    )

    for i in incidencias_qs:
        incidencias.append(i.tipo if hasattr(i, "tipo") else "INCIDENCIA")

    return incidencias

@solo_operativo
def reporte_asistencia(request):
    from nucleo.models import Empleado
    


    empresa = request.empresa

    print("EMPRESA ACTUAL:", request.empresa)


    registros = obtener_asistencias_base(request)

    
    print("TOTAL REGISTROS:", registros.count())
    registros = aplicar_filtros_asistencia(request, registros)
    registros = registros.order_by("empleado__numero_empleado", "-fecha")

    empleados = Empleado.objects.filter(
        empresa=empresa,
        activo=True
    )

    

    registros = list(registros)  # 🔥 FORZAR evaluación


    for r in registros:
        r.estado = calcular_estado_asistencia(r.empleado, r.fecha)
        r.incidencias = calcular_incidencias_asistencia(r.empleado, r.fecha) 
         
       
    return render(request, "reportes/asistencia.html", {
        "resultados": registros,
        "empleados": empleados,
        "inicio": request.GET.get("inicio"),
        "fin": request.GET.get("fin"),
        "empleado_id": request.GET.get("empleado"),
    })




def obtener_asistencias_base(request):
    from asistencia.models import Asistencia

    empresa = request.empresa

    return Asistencia.objects.filter(
        empleado__empresa=empresa
    ).select_related("empleado")

def aplicar_filtros_asistencia(request, queryset):
    empleado_id = request.GET.get("empleado")
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")

    if empleado_id and empleado_id not in ["", "0"]:
        queryset = queryset.filter(empleado_id=empleado_id)

    if inicio:
        queryset = queryset.filter(fecha__gte=inicio)

    if fin:
        queryset = queryset.filter(fecha__lte=fin)

    return queryset



def obtener_movimientos(empresa, fecha_inicio=None, fecha_fin=None, empleado=None):

    qs = Movimiento.objects.filter(asistencia__empresa=empresa)

    # 🔹 usar asistencia__fecha (CORRECTO)
    if fecha_inicio and fecha_fin and fecha_inicio == fecha_fin:
        qs = qs.filter(asistencia__fecha=fecha_inicio)
    else:
        if fecha_inicio:
            qs = qs.filter(asistencia__fecha__gte=fecha_inicio)
        if fecha_fin:
            qs = qs.filter(asistencia__fecha__lte=fecha_fin)

    # 🔹 empleado
    if empleado:
        qs = qs.filter(asistencia__empleado=empleado)

    return qs.select_related("asistencia", "asistencia__empleado").order_by("hora")

@solo_operativo
def reporte_permisos(request):
    from nucleo.models import Empleado

    empresa = request.empresa

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    empleado_id = request.GET.get("empleado")

    resultados = obtener_resultados_permisos(
        empresa=empresa,
        inicio=inicio,
        fin=fin,
        empleado_id=empleado_id,
    )

    empleados = Empleado.objects.filter(
        empresa=empresa,
        activo=True,
    ).order_by(
        "numero_empleado",
        "nombre",
    )

    return render(
        request,
        "reportes/permisos.html",
        {
            "resultados": resultados,
            "empleados": empleados,
            "empleado_id": empleado_id,
            "inicio": inicio,
            "fin": fin,
        },
    )
    
def obtener_asistencias(empresa, inicio=None, fin=None, empleado_id=None):

    asistencias = Asistencia.objects.filter(
        empleado__empresa=empresa
    ).select_related("empleado")

    # 🔹 filtros opcionales
    if inicio:
        asistencias = asistencias.filter(fecha__gte=inicio)

    if fin:
        asistencias = asistencias.filter(fecha__lte=fin)

    if empleado_id:
        asistencias = asistencias.filter(empleado_id=empleado_id)

    # 🔥 calcular estado
    for asistencia in asistencias:
        asistencia.estado = calcular_estado_asistencia(
            asistencia.empleado,
            asistencia.fecha
        )

    return asistencias

@solo_operativo
def exportar_tiempos_extra_excel(request):

    from core.services.asistencia_service import obtener_tiempo_extra

    empresa = request.empresa

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    empleado_id = request.GET.get("empleado")

    # =========================
    # CONSULTA BASE
    # =========================
    asistencias = Asistencia.objects.filter(
        empleado__empresa=empresa
    ).select_related("empleado")

    if inicio and fin:
        asistencias = asistencias.filter(
            fecha__range=(inicio, fin)
        )
    elif inicio:
        asistencias = asistencias.filter(
            fecha__gte=inicio
        )
    elif fin:
        asistencias = asistencias.filter(
            fecha__lte=fin
        )

    if empleado_id and empleado_id != "0":
        asistencias = asistencias.filter(
            empleado_id=empleado_id
        )

    asistencias = asistencias.order_by(
        "empleado__numero_empleado",
        "fecha",
    )

    # =========================
    # CREAR EXCEL
    # =========================
    wb = Workbook()
    ws = wb.active
    ws.title = "Tiempos Extra"

    fila_encabezado = 6
    ultima_columna = "F"

    escribir_encabezado_reporte(
        ws=ws,
        titulo="REPORTE DE TIEMPOS EXTRA",
        empresa=empresa,
        inicio=inicio,
        fin=fin,
        ultima_columna=ultima_columna,
    )

    encabezados = [
        "No. Empleado",
        "Empleado",
        "Fecha",
        "Hora Inicio",
        "Hora Fin",
        "Horas",
    ]

    escribir_encabezados(
        ws=ws,
        fila=fila_encabezado,
        encabezados=encabezados,
    )

    # =========================
    # DATOS
    # =========================
    fila = fila_encabezado + 1
    total_horas = 0
    total_registros = 0

    for asistencia in asistencias:

        info = obtener_tiempo_extra(asistencia)

        if not info:
            continue

        horas = info.get("horas", 0)

        # No mostrar registros cerrados con cero horas
        info = obtener_tiempo_extra(asistencia)

        if not info:
            continue

        horas = info.get("horas", 0)

        valores = [
            asistencia.empleado.numero_empleado,
            asistencia.empleado.nombre,
            asistencia.fecha,
            info.get("inicio") or "--",
            info.get("fin") or "--",
            horas,
        ]

        valores = [
            asistencia.empleado.numero_empleado,
            asistencia.empleado.nombre,
            asistencia.fecha,
            info.get("inicio") or "--",
            info.get("fin") or "--",
            horas,
        ]

        for columna, valor in enumerate(valores, start=1):
            celda = ws.cell(
                row=fila,
                column=columna,
                value=valor,
            )
            celda.border = BORDE_FINO

        ws.cell(fila, 1).alignment = ALINEACION_CENTRO
        ws.cell(fila, 2).alignment = ALINEACION_IZQUIERDA

        for columna in [3, 4, 5, 6]:
            ws.cell(fila, columna).alignment = ALINEACION_CENTRO

        ws.cell(fila, 3).number_format = "dd/mm/yyyy"

        if info.get("inicio"):
            ws.cell(fila, 4).number_format = "hh:mm"

        if info.get("fin"):
            ws.cell(fila, 5).number_format = "hh:mm"

        horas_celda = ws.cell(fila, 6)

        if isinstance(horas, (int, float)):
            horas_celda.fill = RELLENO_OK
            total_horas += horas
        else:
            horas_celda.fill = RELLENO_ALERTA

        if not info.get("fin"):
            ws.cell(fila, 5).fill = RELLENO_ERROR

        fila += 1
        total_registros += 1

    # =========================
    # TOTALES
    # =========================
    fila_total = fila + 1

    ws.merge_cells(
        start_row=fila_total,
        start_column=1,
        end_row=fila_total,
        end_column=4,
    )

    ws.cell(
        row=fila_total,
        column=1,
        value=f"TOTAL DE REGISTROS: {total_registros}",
    )

    ws.cell(
        row=fila_total,
        column=5,
        value="TOTAL HORAS",
    )

    ws.cell(
        row=fila_total,
        column=6,
        value=total_horas,
    )

    for columna in range(1, 7):
        celda = ws.cell(
            row=fila_total,
            column=columna,
        )
        celda.font = FUENTE_NEGRITA
        celda.fill = RELLENO_GRIS
        celda.border = BORDE_FINO

    ws.cell(fila_total, 1).alignment = ALINEACION_DERECHA
    ws.cell(fila_total, 5).alignment = ALINEACION_DERECHA
    ws.cell(fila_total, 6).alignment = ALINEACION_CENTRO

    # =========================
    # ANCHOS
    # =========================
    anchos = {
        "A": 14,
        "B": 35,
        "C": 13,
        "D": 14,
        "E": 14,
        "F": 12,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    # =========================
    # IMPRESIÓN Y RESPUESTA
    # =========================
    configurar_impresion(
        ws=ws,
        fila_encabezado=fila_encabezado,
        ultima_columna=ultima_columna,
        ultima_fila=fila_total,
    )

    return crear_respuesta_excel(
        workbook=wb,
        nombre_archivo="reporte_tiempos_extra.xlsx",
    )


@solo_operativo
def exportar_movimientos_excel(request):

    empresa = request.empresa

    fecha_inicio = request.GET.get("inicio")
    fecha_fin = request.GET.get("fin")
    empleado_id = request.GET.get("empleado")

    if fecha_inicio == "None":
        fecha_inicio = None

    if fecha_fin == "None":
        fecha_fin = None

    if empleado_id == "None":
        empleado_id = None

    empleado = None

    if empleado_id:
        try:
            empleado = Empleado.objects.get(
                id=empleado_id,
                empresa=empresa,
            )
        except Empleado.DoesNotExist:
            empleado = None

    movimientos = obtener_movimientos(
        empresa=empresa,
        fecha_inicio=fecha_inicio,
        fecha_fin=fecha_fin,
        empleado=empleado,
    )

    if empleado_id:
        movimientos = movimientos.filter(
            asistencia__empleado_id=int(empleado_id)
        )

    movimientos = movimientos.order_by(
        "asistencia__empleado__numero_empleado",
        "fecha",
        "hora",
    )

    # Crear libro antes de usar ws
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    fila_encabezado = 6
    ultima_columna = "E"

    escribir_encabezado_reporte(
        ws=ws,
        titulo="REPORTE DE MOVIMIENTOS",
        empresa=empresa,
        inicio=fecha_inicio,
        fin=fecha_fin,
        ultima_columna=ultima_columna,
    )

    encabezados = [
        "No. Empleado",
        "Empleado",
        "Fecha",
        "Hora",
        "Movimiento",
    ]

    escribir_encabezados(
        ws=ws,
        fila=fila_encabezado,
        encabezados=encabezados,
    )

    fila = fila_encabezado + 1
    total_registros = 0

    for movimiento in movimientos:

        valores = [
            movimiento.asistencia.empleado.numero_empleado,
            movimiento.asistencia.empleado.nombre,
            movimiento.fecha,
            movimiento.hora,
            movimiento.get_tipo_display(),
        ]

        for columna, valor in enumerate(valores, start=1):
            celda = ws.cell(
                row=fila,
                column=columna,
                value=valor,
            )
            celda.border = BORDE_FINO

        ws.cell(fila, 1).alignment = ALINEACION_CENTRO
        ws.cell(fila, 2).alignment = ALINEACION_IZQUIERDA
        ws.cell(fila, 3).alignment = ALINEACION_CENTRO
        ws.cell(fila, 4).alignment = ALINEACION_CENTRO
        ws.cell(fila, 5).alignment = ALINEACION_CENTRO

        movimiento_celda = ws.cell(fila, 5)
        tipo_movimiento = movimiento.tipo.strip().upper()

        if tipo_movimiento == "ENTRADA":
            movimiento_celda.fill = RELLENO_OK

        elif tipo_movimiento == "SALIDA":
            movimiento_celda.fill = RELLENO_GRIS

        elif tipo_movimiento == "SALIDA_PERMISO":
            movimiento_celda.fill = RELLENO_ALERTA

        elif tipo_movimiento == "REGRESO":
            movimiento_celda.fill = RELLENO_INFORMATIVO

        elif tipo_movimiento == "INICIO_TIEMPO_EXTRA":
            movimiento_celda.fill = RELLENO_ALERTA

        elif tipo_movimiento == "FIN_TIEMPO_EXTRA":
            movimiento_celda.fill = RELLENO_OK

        ws.cell(fila, 3).number_format = "dd/mm/yyyy"
        ws.cell(fila, 4).number_format = "hh:mm"

        fila += 1
        total_registros += 1

    fila_total = fila + 1

    ws.merge_cells(
        start_row=fila_total,
        start_column=1,
        end_row=fila_total,
        end_column=4,
    )

    ws.cell(
        row=fila_total,
        column=1,
        value="TOTAL DE REGISTROS",
    )

    ws.cell(
        row=fila_total,
        column=5,
        value=total_registros,
    )

    for columna in range(1, 6):
        celda = ws.cell(
            row=fila_total,
            column=columna,
        )
        celda.font = FUENTE_NEGRITA
        celda.fill = RELLENO_GRIS
        celda.border = BORDE_FINO

    ws.cell(fila_total, 1).alignment = ALINEACION_DERECHA
    ws.cell(fila_total, 5).alignment = ALINEACION_CENTRO

    anchos = {
        "A": 14,
        "B": 35,
        "C": 13,
        "D": 12,
        "E": 28,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    configurar_impresion(
        ws=ws,
        fila_encabezado=fila_encabezado,
        ultima_columna=ultima_columna,
        ultima_fila=fila_total,
    )

    return crear_respuesta_excel(
        workbook=wb,
        nombre_archivo="reporte_movimientos.xlsx",
    )

  
 
def reporte_excel(request):
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    

    registros = obtener_asistencias_base(request)
    registros = aplicar_filtros_asistencia(request, registros)
    registros = registros.order_by(
    "empleado__numero_empleado",
    "-fecha",
)

    response = HttpResponse(content_type="text/csv")

    # ✅ nombre dinámico
    fecha = datetime.now().strftime("%Y%m%d_%H%M")
    import uuid

    response["Content-Disposition"] = f'attachment; filename="asistencia_{uuid.uuid4().hex}.csv"'

    writer = csv.writer(response)

    # ✅ encabezados claros
    writer.writerow([
        "Empleado",
        "Fecha",
        "Hora Entrada",
        "Hora Salida",
        "Estado",
        "Incidencias"
    ])

    for r in registros:
        estado = calcular_estado_asistencia(r.empleado, r.fecha)
        incidencias = calcular_incidencias_asistencia(r.empleado, r.fecha)

        writer.writerow([
            r.empleado.nombre,
            r.fecha.strftime("%d/%m/%Y"),
            r.hora_entrada.strftime("%H:%M") if r.hora_entrada else "--",
            r.hora_salida.strftime("%H:%M") if r.hora_salida else "--",
            estado,
            " | ".join(incidencias) if incidencias else "OK"
        ])
    return response

@solo_operativo
def reporte_excel_xlsx(request):

    configuracion = construir_reporte_asistencia(request)

    return crear_reporte_excel(
        **configuracion
    )
    

   
@solo_operativo
def reporte_tiempos_extra(request):
    

    from asistencia.models import Asistencia, Movimiento
    from nucleo.models import Empleado
    from datetime import datetime

    empresa = request.empresa

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    empleado_id = request.GET.get("empleado")

    asistencias = Asistencia.objects.filter(
        empleado__empresa=empresa
    ).select_related("empleado")

    # 🔹 filtros
    if inicio and fin:
        asistencias = asistencias.filter(fecha__range=[inicio, fin])
    elif inicio:
        asistencias = asistencias.filter(fecha__gte=inicio)
    elif fin:
        asistencias = asistencias.filter(fecha__lte=fin)

    if empleado_id and empleado_id != "":
        asistencias = asistencias.filter(empleado_id=empleado_id)

    asistencias = asistencias.order_by(
        "empleado__numero_empleado",
        "fecha"
    )

    resultados = []
    total_horas = 0

    for a in asistencias:

        movimientos = Movimiento.objects.filter(
            asistencia__empleado=a.empleado,
            asistencia__fecha=a.fecha
        ).order_by("hora")

        inicio_extra = None
        fin_extra = None

        for m in movimientos:
            if m.tipo == "INICIO_TIEMPO_EXTRA":
                inicio_extra = m.hora
            elif m.tipo == "FIN_TIEMPO_EXTRA":
                fin_extra = m.hora

        if inicio_extra:

            if fin_extra:
                # 🔥 calcular minutos totales
                t_inicio = datetime.combine(a.fecha, inicio_extra)
                t_fin = datetime.combine(a.fecha, fin_extra)

                diff = t_fin - t_inicio

                total_minutos = int(diff.total_seconds() / 60)

                horas_base = total_minutos // 60
                minutos = total_minutos % 60

                # 🔥 REGLA DE NEGOCIO
                if minutos >= 45:
                    horas_final = horas_base + 1
                else:
                    horas_final = horas_base

                total_horas += horas_final

                resultados.append({
                    "empleado": a.empleado,
                    "fecha": a.fecha,
                    "hora_inicio": inicio_extra,
                    "hora_fin": fin_extra,
                    "horas": horas_final   # 🔥 SOLO ENTERO
                })

            else:
                # 🔥 EN CURSO
                resultados.append({
                    "empleado": a.empleado,
                    "fecha": a.fecha,
                    "hora_inicio": inicio_extra,
                    "hora_fin": None,
                    "horas": "EN CURSO"
                })
                
    empleados = Empleado.objects.filter(
        empresa=empresa,
        activo=True
    )

    context = {
        "tiempos": resultados,
        "empleados": empleados,
        "inicio": inicio,
        "fin": fin,
        "empleado_id": empleado_id,
        "total_horas": round(total_horas, 2)
    }

    return render(request, "reportes/tiempo_extra.html", context)

@solo_operativo
def reporte_movimientos(request):

    empresa = request.empresa

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    empleado_id = request.GET.get("empleado")

    # 🔥 limpiar valores
    if inicio == "None":
        inicio = None
    if fin == "None":
        fin = None
    if empleado_id == "None":
        empleado_id = None

    movimientos = Movimiento.objects.select_related(
        "asistencia__empleado"
    ).filter(
        asistencia__empleado__empresa=empresa
    )

    # 🔥 FILTROS
    if inicio and fin:
        movimientos = movimientos.filter(fecha__range=[inicio, fin])
    elif inicio:
        movimientos = movimientos.filter(fecha__gte=inicio)
    elif fin:
        movimientos = movimientos.filter(fecha__lte=fin)

    if empleado_id:
        movimientos = movimientos.filter(
            asistencia__empleado_id=int(empleado_id)
        )

    movimientos = movimientos.order_by(
        "asistencia__empleado", "fecha", "hora"
    )

    empleados = Empleado.objects.filter(
        empresa=empresa,
        activo=True
    )

    return render(request, "reportes/movimientos.html", {
        "movimientos": movimientos,
        "empleados": empleados,
        "empleado_id": empleado_id,
        "inicio": inicio,
        "fin": fin
    })

    
@solo_operativo
def reporte_incidencias(request):

    empresa = request.empresa

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    empleado_id = request.GET.get("empleado")

    # ✅ PRIMERO crear queryset base
    incidencias = Incidencia.objects.filter(
        empleado__empresa=empresa
    ).select_related("empleado")

    # ✅ luego aplicar filtros
    if inicio and fin:
        incidencias = incidencias.filter(fecha_inicio__range=[inicio, fin])

    elif inicio:
        incidencias = incidencias.filter(fecha_inicio__gte=inicio)

    elif fin:
        incidencias = incidencias.filter(fecha_inicio__lte=fin)

    if empleado_id and empleado_id != "0":
        incidencias = incidencias.filter(empleado_id=empleado_id)

    empleados = Empleado.objects.filter(
        empresa=empresa,
        activo=True
    )
    context = {
        "incidencias": incidencias,
        "empleados": empleados,
        "inicio": inicio,
        "fin": fin,
        "empleado_id": empleado_id
    }

    return render(request, "reportes/incidencias.html", context)

    



def index(request):
    return render(request, "reportes/index.html")


   
@solo_operativo
def exportar_incidencias_excel_xlsx(request):

    empresa = request.empresa

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    empleado_id = request.GET.get("empleado")

    incidencias = (
        Incidencia.objects.filter(
            empleado__empresa=empresa
        )
        .select_related("empleado")
        .order_by(
            "empleado__numero_empleado",
            "-fecha_inicio"
        )
    )

    if empleado_id:
        incidencias = incidencias.filter(
            empleado_id=empleado_id
        )

    if inicio:
        incidencias = incidencias.filter(
            fecha_inicio__gte=inicio
        )

    if fin:
        incidencias = incidencias.filter(
            fecha_inicio__lte=fin
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Incidencias"

    fila_encabezado = 6
    ultima_columna = "E"

    escribir_encabezado_reporte(
        ws=ws,
        titulo="REPORTE DE INCIDENCIAS",
        empresa=empresa,
        inicio=inicio,
        fin=fin,
        ultima_columna=ultima_columna,
    )

    encabezados = [
        "No. Empleado",
        "Empleado",
        "Tipo",
        "Fecha Inicio",
        "Fecha Fin",
    ]

    escribir_encabezados(
        ws=ws,
        fila=fila_encabezado,
        encabezados=encabezados,
    )

    fila = fila_encabezado + 1

    for incidencia in incidencias:

        valores = [
            incidencia.empleado.numero_empleado,
            incidencia.empleado.nombre,
            incidencia.tipo,
            incidencia.fecha_inicio,
            incidencia.fecha_fin,
        ]

        for columna, valor in enumerate(valores, start=1):

            celda = ws.cell(
                row=fila,
                column=columna,
                value=valor,
            )

            celda.border = BORDE_FINO

        ws.cell(fila, 1).alignment = ALINEACION_CENTRO
        ws.cell(fila, 2).alignment = ALINEACION_IZQUIERDA
        ws.cell(fila, 3).alignment = ALINEACION_CENTRO
        ws.cell(fila, 4).alignment = ALINEACION_CENTRO
        ws.cell(fila, 5).alignment = ALINEACION_CENTRO

        ws.cell(fila, 4).number_format = "dd/mm/yyyy"
        ws.cell(fila, 5).number_format = "dd/mm/yyyy"

        tipo = str(incidencia.tipo).upper()

        tipo_cell = ws.cell(fila, 3)

        if tipo == "VACACIONES":
            tipo_cell.fill = RELLENO_INFORMATIVO

        elif tipo == "INCAPACIDAD":
            tipo_cell.fill = RELLENO_ERROR

        elif tipo == "DESCANSO":
            tipo_cell.fill = RELLENO_GRIS

        elif tipo == "PERMISO":
            tipo_cell.fill = RELLENO_ALERTA

        fila += 1

    fila_total = fila + 1

    ws.merge_cells(
        start_row=fila_total,
        start_column=1,
        end_row=fila_total,
        end_column=4,
    )

    ws.cell(
        row=fila_total,
        column=1,
        value="TOTAL DE INCIDENCIAS",
    )

    ws.cell(
        row=fila_total,
        column=5,
        value=incidencias.count(),
    )

    for columna in range(1, 6):

        celda = ws.cell(
            row=fila_total,
            column=columna,
        )

        celda.font = FUENTE_NEGRITA
        celda.fill = RELLENO_GRIS
        celda.border = BORDE_FINO

    ws.cell(fila_total, 1).alignment = ALINEACION_DERECHA
    ws.cell(fila_total, 5).alignment = ALINEACION_CENTRO

    anchos = {
        "A": 14,
        "B": 35,
        "C": 18,
        "D": 14,
        "E": 14,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    configurar_impresion(
        ws=ws,
        fila_encabezado=fila_encabezado,
        ultima_columna=ultima_columna,
        ultima_fila=fila_total,
    )

    return crear_respuesta_excel(
        workbook=wb,
        nombre_archivo="reporte_incidencias.xlsx",
    )
@solo_operativo
def exportar_permisos_excel(request):

    configuracion = construir_reporte_permisos(
        request
    )

    return crear_reporte_excel(
        **configuracion
    )

    # ----- CÓDIGO ANTERIOR ABAJO -----

@solo_operativo
def exportar_permisos_excel(request):

    from asistencia.models import Movimiento

    empresa = request.empresa

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    empleado_id = request.GET.get("empleado")

    # =========================
    # CONSULTA BASE
    # =========================
    movimientos = Movimiento.objects.filter(
        tipo__in=["SALIDA_PERMISO", "REGRESO"],
        asistencia__empleado__empresa=empresa,
    ).select_related(
        "asistencia__empleado"
    )

    if inicio:
        movimientos = movimientos.filter(fecha__gte=inicio)

    if fin:
        movimientos = movimientos.filter(fecha__lte=fin)

    if empleado_id:
        movimientos = movimientos.filter(
            asistencia__empleado_id=empleado_id
        )

    movimientos = movimientos.order_by(
        "asistencia__empleado__numero_empleado",
        "fecha",
        "hora",
    )

    # =========================
    # EMPAREJAR SALIDA Y REGRESO
    # =========================
    control = {}
    resultados = []

    for movimiento in movimientos:

        llave = (
            movimiento.asistencia_id,
            movimiento.fecha,
        )

        if movimiento.tipo == "SALIDA_PERMISO":
            control[llave] = {
                "numero_empleado": (
                    movimiento.asistencia.empleado.numero_empleado
                ),
                "empleado": movimiento.asistencia.empleado.nombre,
                "fecha": movimiento.fecha,
                "salida": movimiento.hora,
                "regreso": None,
            }

        elif movimiento.tipo == "REGRESO" and llave in control:
            datos = control[llave]
            datos["regreso"] = movimiento.hora

            resultados.append(datos)
            del control[llave]

    # Incluir permisos que todavía no tienen regreso
    resultados.extend(control.values())

    resultados.sort(
        key=lambda registro: (
            str(registro["numero_empleado"]),
            registro["fecha"],
            registro["salida"],
        )
    )

    # =========================
    # CREAR EXCEL
    # =========================
    wb = Workbook()
    ws = wb.active
    ws.title = "Permisos"

    fila_encabezado = 6
    ultima_columna = "E"

    escribir_encabezado_reporte(
        ws=ws,
        titulo="REPORTE DE PERMISOS",
        empresa=empresa,
        inicio=inicio,
        fin=fin,
        ultima_columna=ultima_columna,
    )

    encabezados = [
        "No. Empleado",
        "Empleado",
        "Fecha",
        "Salida",
        "Regreso",
    ]

    escribir_encabezados(
        ws=ws,
        fila=fila_encabezado,
        encabezados=encabezados,
    )

    # =========================
    # DATOS
    # =========================
    fila = fila_encabezado + 1

    for registro in resultados:

        valores = [
            registro["numero_empleado"],
            registro["empleado"],
            registro["fecha"],
            registro["salida"] or "--",
            registro["regreso"] or "--",
        ]

        for columna, valor in enumerate(valores, start=1):
            celda = ws.cell(
                row=fila,
                column=columna,
                value=valor,
            )
            celda.border = BORDE_FINO

        ws.cell(fila, 1).alignment = ALINEACION_CENTRO
        ws.cell(fila, 2).alignment = ALINEACION_IZQUIERDA
        ws.cell(fila, 3).alignment = ALINEACION_CENTRO
        ws.cell(fila, 4).alignment = ALINEACION_CENTRO
        ws.cell(fila, 5).alignment = ALINEACION_CENTRO

        ws.cell(fila, 3).number_format = "dd/mm/yyyy"

        if registro["salida"]:
            ws.cell(fila, 4).number_format = "hh:mm"

        if registro["regreso"]:
            ws.cell(fila, 5).number_format = "hh:mm"

        # Colores
        ws.cell(fila, 4).fill = RELLENO_ALERTA

        if registro["regreso"]:
            ws.cell(fila, 5).fill = RELLENO_INFORMATIVO
        else:
            ws.cell(fila, 5).fill = RELLENO_ERROR

        fila += 1

    # =========================
    # TOTAL
    # =========================
    fila_total = fila + 1

    ws.merge_cells(
        start_row=fila_total,
        start_column=1,
        end_row=fila_total,
        end_column=4,
    )

    ws.cell(
        row=fila_total,
        column=1,
        value="TOTAL DE PERMISOS",
    )

    ws.cell(
        row=fila_total,
        column=5,
        value=len(resultados),
    )

    for columna in range(1, 6):
        celda = ws.cell(
            row=fila_total,
            column=columna,
        )
        celda.font = FUENTE_NEGRITA
        celda.fill = RELLENO_GRIS
        celda.border = BORDE_FINO

    ws.cell(fila_total, 1).alignment = ALINEACION_DERECHA
    ws.cell(fila_total, 5).alignment = ALINEACION_CENTRO

    # =========================
    # ANCHOS
    # =========================
    anchos = {
        "A": 14,
        "B": 35,
        "C": 13,
        "D": 12,
        "E": 12,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    # =========================
    # IMPRESIÓN Y RESPUESTA
    # =========================
    configurar_impresion(
        ws=ws,
        fila_encabezado=fila_encabezado,
        ultima_columna=ultima_columna,
        ultima_fila=fila_total,
    )

    return crear_respuesta_excel(
        workbook=wb,
        nombre_archivo="reporte_permisos.xlsx",
    )


def reporte_nomina(request):

    empresa = request.empresa

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")

    asistencias = Asistencia.objects.filter(
        empleado__empresa=empresa
    ).select_related("empleado")

    if inicio and fin:
        asistencias = asistencias.filter(fecha__range=(inicio, fin))
    elif inicio:
        asistencias = asistencias.filter(fecha__gte=inicio)
    elif fin:
        asistencias = asistencias.filter(fecha__lte=fin)

    empleados = {}
    total_general = 0  # 🔥 SIEMPRE DEFINIDO

    for a in asistencias:
        emp = a.empleado

        if emp.id not in empleados:
            empleados[emp.id] = {
                "empleado": emp,
                "horas": 0,
                "pago": 0
            }

        info = obtener_tiempo_extra(a)  # 🔥 ESTA LÍNEA FALTABA

        horas = 0

        if info:
            horas = float(info.get("horas", 0))

        pago = horas * float(emp.costo_hora or 0) * 2

        empleados[emp.id]["horas"] += horas
        empleados[emp.id]["pago"] += pago
        
        

    # 🔥 SIEMPRE SE EJECUTA (aunque no haya datos)
    total_general = sum(e["pago"] for e in empleados.values()) if empleados else 0
    context = {
        "empleados": empleados.values(),
        "inicio": inicio,
        "fin": fin,
        "total_general": total_general
}
    return render(request, "reportes/nomina.html", context)

    from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from core.services.asistencia_service import obtener_tiempo_extra


def exportar_nomina_excel(request):

    empresa = request.empresa

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")

    asistencias = Asistencia.objects.filter(
        empleado__empresa=empresa
    ).select_related("empleado")

    if inicio and fin:
        asistencias = asistencias.filter(fecha__range=(inicio, fin))

    empleados = {}

    for a in asistencias:
        emp = a.empleado

        if emp.id not in empleados:
            empleados[emp.id] = {
                "empleado": emp,
                "horas": 0,
                "pago": 0
            }

        info = obtener_tiempo_extra(a)

        if info and isinstance(info["horas"], int):
            horas = info["horas"]
            pago = horas * float(emp.costo_hora) * 2

            empleados[emp.id]["horas"] += horas
            empleados[emp.id]["pago"] += pago

    # 🔥 CREAR EXCEL
    wb = Workbook()
    ws = wb.active
    ws.title = "Nómina"

    # 🔹 ENCABEZADOS
    encabezados = ["No. Empleado", "Empleado", "Horas", "Costo Hora", "Pago"]

    for col, valor in enumerate(encabezados, 1):
        cell = ws.cell(row=1, column=col, value=valor)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    fila = 2
    total_general = 0

    for e in empleados.values():

        if e["horas"] > 0:

            ws.cell(row=fila, column=1, value=e["empleado"].numero_empleado)
            ws.cell(row=fila, column=2, value=e["empleado"].nombre)
            ws.cell(row=fila, column=3, value=e["horas"])
            ws.cell(row=fila, column=4, value=float(e["empleado"].costo_hora))
            ws.cell(row=fila, column=5, value=e["pago"])

            total_general += e["pago"]
            fila += 1

    # 🔹 TOTAL
    ws.cell(row=fila + 1, column=4, value="TOTAL").font = Font(bold=True)
    ws.cell(row=fila + 1, column=5, value=total_general).font = Font(bold=True)

    # 🔹 AUTO AJUSTE
    for column in ws.columns:
        max_length = 0
        col_letter = column[0].column_letter

        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2

    # 🔹 RESPONSE
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=nomina.xlsx"

    wb.save(response)
    return response

@solo_operativo
def reporte_pre_nomina(request):
    empresa = request.empresa

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    departamento_id = request.GET.get("departamento")
    empleado_id = request.GET.get("empleado")

    resultados = []

    if inicio and fin:
        resultados = ResumenPrenomina(
            empresa=empresa,
            fecha_inicio=inicio,
            fecha_fin=fin,
            departamento_id=departamento_id or None,
            empleado_id=empleado_id or None,
        ).generar()

    departamentos = Departamento.objects.filter(
        empresa=empresa,
        activo=True,
    ).order_by("nombre")

    empleados = Empleado.objects.filter(
        empresa=empresa,
        activo=True,
    ).order_by("numero_empleado")

    return render(request, "reportes/pre_nomina.html", {
        "empresa": empresa,
        "resultados": resultados,
        "departamentos": departamentos,
        "empleados": empleados,
        "inicio": inicio,
        "fin": fin,
        "departamento_id": departamento_id,
        "empleado_id": empleado_id,
    })

@solo_operativo
def exportar_pre_nomina_excel(request):

    empresa = request.empresa

    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    departamento_id = request.GET.get("departamento")
    empleado_id = request.GET.get("empleado")

    if not inicio or not fin:
        messages.error(
            request,
            "Debe seleccionar fecha de inicio y fecha de fin."
        )
        return redirect("reportes:pre_nomina")

    resultados = ResumenPrenomina(
        empresa=empresa,
        fecha_inicio=inicio,
        fecha_fin=fin,
        departamento_id=departamento_id or None,
        empleado_id=empleado_id or None,
    ).generar()

    # =========================
    # CREAR LIBRO
    # =========================
    wb = Workbook()
    ws = wb.active
    ws.title = "Pre-Nómina"

    fila_encabezado = 6

    # 17 columnas: A hasta R
    ultima_columna = "R"

    escribir_encabezado_reporte(
        ws=ws,
        titulo="RESUMEN SEMANAL DE PRE-NÓMINA",
        empresa=empresa,
        inicio=inicio,
        fin=fin,
        ultima_columna=ultima_columna,
    )

    # =========================
    # ENCABEZADOS
    # =========================
    encabezados = [
        "No. Empleado",
        "Empleado",
        "Departamento",
        "Turno",

        # Cuadre del período
        "Días Trabajados",
        "Días de Descanso",
        "Faltas",
        "Vacaciones",
        "Incapacidades",
        "Licencias",
        "Permisos",
        "Días a Pagar",
        "Días Clasificados",
        "Período",

        # Información adicional
        "Retardos",
        "Salidas Permiso",
        "Horas Trabajadas",
        "Tiempo Extra",
    ]

    escribir_encabezados(
        ws=ws,
        fila=fila_encabezado,
        encabezados=encabezados,
    )

    # =========================
    # TOTALES
    # =========================
    totales = {
        "dias_laborados": 0,
        "no_laborales": 0,
        "faltas": 0,
        "vacaciones": 0,
        "incapacidades": 0,
        "descansos": 0,
        "permisos": 0,
        "dias_a_pagar": 0,
        "total_fila": 0,
        "dias_periodo": 0,
        "retardos": 0,
        "salidas_permiso": 0,
        "horas_trabajadas": 0,
        "tiempo_extra": 0,
    }

    # =========================
    # DATOS
    # =========================
    fila = fila_encabezado + 1

    for registro in resultados:

        valores = [
            registro["numero_empleado"],
            registro["empleado"],
            registro["departamento"],
            registro["turno"],

            registro["dias_laborados"],
            registro["no_laborales"],
            registro["faltas"],
            registro["vacaciones"],
            registro["incapacidades"],
            registro["descansos"],     # visible como Licencias
            registro["permisos"],
            registro["dias_a_pagar"],
            registro["total_fila"],
            registro["dias_periodo"],

            registro["retardos"],
            registro["salidas_permiso"],
            registro["horas_trabajadas"],
            registro["tiempo_extra"],
        ]

        for columna, valor in enumerate(valores, start=1):

            celda = ws.cell(
                row=fila,
                column=columna,
                value=valor,
            )

            celda.border = BORDE_FINO

        # =========================
        # ALINEACIÓN
        # =========================

        # Número de empleado y turno
        for columna in [1, 4]:
            ws.cell(
                row=fila,
                column=columna,
            ).alignment = ALINEACION_CENTRO

        # Empleado y departamento
        for columna in [2, 3]:
            ws.cell(
                row=fila,
                column=columna,
            ).alignment = ALINEACION_IZQUIERDA

        # Valores numéricos
        for columna in range(5, 18):
            ws.cell(
                row=fila,
                column=columna,
            ).alignment = ALINEACION_CENTRO

        # Horas trabajadas y tiempo extra
        ws.cell(fila, 17).number_format = "0.00"
        ws.cell(fila, 18).number_format = "0.00"

        # =========================
        # COLORES
        # =========================

        if registro["dias_a_pagar"] > 0:
            ws.cell(fila, 12).fill = RELLENO_OK

        if registro["faltas"] > 0:
            ws.cell(fila, 7).fill = RELLENO_ERROR

        if registro["vacaciones"] > 0:
            ws.cell(fila, 8).fill = RELLENO_INFORMATIVO

        if registro["incapacidades"] > 0:
            ws.cell(fila, 9).fill = RELLENO_ERROR

        if registro["descansos"] > 0:
            ws.cell(fila, 10).fill = RELLENO_GRIS

        if registro["permisos"] > 0:
            ws.cell(fila, 11).fill = RELLENO_ALERTA

        if registro.get("diferencia", 0) == 0:
            ws.cell(fila, 13).fill = RELLENO_OK
        else:
            ws.cell(fila, 13).fill = RELLENO_ERROR

        ws.cell(fila, 14).fill = RELLENO_GRIS

        if registro["retardos"] > 0:
            ws.cell(fila, 15).fill = RELLENO_ALERTA

        if registro["salidas_permiso"] > 0:
            ws.cell(fila, 16).fill = RELLENO_INFORMATIVO

        if registro["tiempo_extra"] > 0:
            ws.cell(fila, 18).fill = RELLENO_OK
        # =========================
        # ACUMULAR TOTALES
        # =========================
        for clave in totales:

            valor = registro.get(clave, 0) or 0

            if isinstance(valor, (int, float)):
                totales[clave] += valor

        fila += 1

    # =========================
    # FILA DE TOTALES
    # =========================
    fila_total = fila + 1

    ws.merge_cells(
        start_row=fila_total,
        start_column=1,
        end_row=fila_total,
        end_column=4,
    )

    ws.cell(
        row=fila_total,
        column=1,
        value="TOTALES GENERALES",
    )

    valores_totales = [
        totales["dias_laborados"],
        totales["no_laborales"],
        totales["faltas"],
        totales["vacaciones"],
        totales["incapacidades"],
        totales["descansos"],
        totales["permisos"],
        totales["dias_a_pagar"],
        totales["total_fila"],
        totales["dias_periodo"],
        totales["retardos"],
        totales["salidas_permiso"],
        round(totales["horas_trabajadas"], 2),
        round(totales["tiempo_extra"], 2),
    ]

    for columna, valor in enumerate(
        valores_totales,
        start=5,
    ):
        ws.cell(
            row=fila_total,
            column=columna,
            value=valor,
        )

    for columna in range(1, 19):

        celda = ws.cell(
            row=fila_total,
            column=columna,
        )

        celda.font = FUENTE_NEGRITA
        celda.fill = RELLENO_GRIS
        celda.border = BORDE_FINO
        celda.alignment = ALINEACION_CENTRO

    ws.cell(
        row=fila_total,
        column=1,
    ).alignment = ALINEACION_DERECHA

    ws.cell(fila_total, 17).number_format = "0.00"
    ws.cell(fila_total, 18).number_format = "0.00"

    # =========================
    # ANCHOS
    # =========================
    anchos = {
        "A": 14,
        "B": 32,
        "C": 24,
        "D": 20,
        "E": 15,
        "F": 16,
        "G": 9,
        "H": 12,
        "I": 15,
        "J": 11,
        "K": 10,
        "L": 14,
        "M": 17,
        "N": 10,
        "O": 10,
        "P": 16,
        "Q": 17,
        "R": 13,
    }

    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    # =========================
    # IMPRESIÓN
    # =========================
    configurar_impresion(
        ws=ws,
        fila_encabezado=fila_encabezado,
        ultima_columna=ultima_columna,
        ultima_fila=fila_total,
    )

    return crear_respuesta_excel(
        workbook=wb,
        nombre_archivo="resumen_pre_nomina.xlsx",
    )