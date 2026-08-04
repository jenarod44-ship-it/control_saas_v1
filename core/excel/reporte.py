from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from copy import copy

from .encabezado import (
    escribir_encabezado_reporte,
    escribir_encabezados,
)
from .estilos import (
    ALINEACION_CENTRO,
    ALINEACION_IZQUIERDA,
    ALINEACION_ENVUELTA,
    BORDE_FINO,
    RELLENO_ALERTA,
    RELLENO_ERROR,
    RELLENO_GRIS,
    RELLENO_INFORMATIVO,
    RELLENO_OK,
)
from .impresion import configurar_impresion
from .respuesta import crear_respuesta_excel


RELLENOS_ESTADO = {
    "OK": RELLENO_OK,
    "ASISTENCIA": RELLENO_OK,
    "COMPLETO": RELLENO_OK,

    "RETARDO": RELLENO_ALERTA,

    "FALTA": RELLENO_ERROR,

    "INCOMPLETO": RELLENO_INFORMATIVO,
    "PENDIENTE": RELLENO_INFORMATIVO,
    "TIEMPO_EXTRA": RELLENO_INFORMATIVO,

    "VACACIONES": RELLENO_GRIS,
    "INCAPACIDAD": RELLENO_GRIS,
    "DESCANSO": RELLENO_GRIS,
    "PERMISO": RELLENO_GRIS,
    "INCIDENCIA": RELLENO_GRIS,
    "NO_LABORAL": RELLENO_GRIS,
    "SIN CONTROL": RELLENO_GRIS,
    "SIN_TURNO": RELLENO_GRIS,
}


def crear_reporte_excel(
    *,
    titulo,
    empresa,
    nombre_hoja,
    nombre_archivo,
    encabezados,
    filas,
    anchos=None,
    inicio=None,
    fin=None,
    formatos=None,
    columnas_centradas=None,
    columnas_envueltas=None,
    columna_estado=None,
    reglas_color=None,
):
    """
    Genera un reporte Excel con el formato estándar del sistema.

    Parámetros:
    - encabezados: lista de nombres de columnas.
    - filas: lista de listas o tuplas con los valores.
    - anchos: diccionario {1: 10, 2: 25, ...}.
    - formatos: diccionario {1: "0", 4: "hh:mm", ...}.
    - columnas_centradas: números de columna.
    - columnas_envueltas: números de columna.
    - columna_estado: número de columna que contiene el estado.
    """

    wb = Workbook()
    ws = wb.active
    ws.title = nombre_hoja[:31]

    total_columnas = len(encabezados)
    ultima_columna = get_column_letter(total_columnas)

    escribir_encabezado_reporte(
        ws=ws,
        titulo=titulo,
        empresa=empresa,
        inicio=inicio,
        fin=fin,
        ultima_columna=ultima_columna,
    )

    fila_encabezado = 6

    escribir_encabezados(
        ws,
        fila_encabezado,
        encabezados,
    )

    ws.row_dimensions[fila_encabezado].height = 22

    formatos = formatos or {}
    columnas_centradas = set(columnas_centradas or [])
    columnas_envueltas = set(columnas_envueltas or [])
    reglas_color = reglas_color or {}

    fila_actual = fila_encabezado + 1

    for valores in filas:

        for numero_columna, valor in enumerate(valores, start=1):
            celda = ws.cell(
                row=fila_actual,
                column=numero_columna,
                value=valor,
            )

            celda.border = BORDE_FINO

            if numero_columna in columnas_envueltas:
                celda.alignment = ALINEACION_ENVUELTA

            elif numero_columna in columnas_centradas:
                celda.alignment = ALINEACION_CENTRO

            else:
                celda.alignment = ALINEACION_IZQUIERDA

            if numero_columna in formatos and valor not in [None, "--"]:
                celda.number_format = formatos[numero_columna]

                # ==========================================
        # REGLAS GENÉRICAS DE COLOR
        # ==========================================
        for numero_columna, mapa_colores in reglas_color.items():

            if numero_columna < 1 or numero_columna > len(valores):
                continue

            valor = valores[numero_columna - 1]

            valor_normalizado = str(
                valor or ""
            ).strip().upper()

            relleno = mapa_colores.get(
                valor_normalizado
            )

            # Permite reglas parciales como:
            # "SIN SALIDA", "RETARDO" o "SIN TURNO"
            # dentro de textos más largos.
            if not relleno:
                for texto_regla, relleno_regla in mapa_colores.items():
                    texto_normalizado = str(
                        texto_regla
                    ).strip().upper()

                    if (
                        texto_normalizado
                        and texto_normalizado in valor_normalizado
                    ):
                        relleno = relleno_regla
                        break

            if relleno:
                ws.cell(
                    row=fila_actual,
                    column=numero_columna,
                ).fill = relleno

        # ==========================================
        # COMPATIBILIDAD TEMPORAL
        # Estado del Día todavía usa columna_estado.
        # ==========================================
        if columna_estado and columna_estado not in reglas_color:

            estado = valores[columna_estado - 1]

            estado_normalizado = str(
                estado or ""
            ).strip().upper()

            relleno = RELLENOS_ESTADO.get(
                estado_normalizado
            )

            if relleno:
                ws.cell(
                    row=fila_actual,
                    column=columna_estado,
                ).fill = relleno

        fila_actual += 1

    total_registros = len(filas)

    # =========================
    # TOTAL
    # =========================
    fila_total = fila_actual + 1

    if total_columnas > 1:
        ws.merge_cells(
            start_row=fila_total,
            start_column=1,
            end_row=fila_total,
            end_column=total_columnas - 1,
        )

    celda_total_texto = ws.cell(
        row=fila_total,
        column=1,
        value="TOTAL DE REGISTROS",
    )

    celda_total_numero = ws.cell(
        row=fila_total,
        column=total_columnas,
        value=total_registros,
    )

    for numero_columna in range(1, total_columnas + 1):
        celda = ws.cell(
            row=fila_total,
            column=numero_columna,
        )
        celda.fill = RELLENO_GRIS
        celda.border = BORDE_FINO
        fuente = copy(celda.font)
        fuente.bold = True
        celda.font = fuente

    celda_total_texto.alignment = ALINEACION_CENTRO
    celda_total_numero.alignment = ALINEACION_CENTRO

    # =========================
    # ANCHOS
    # =========================
    for numero_columna in range(1, total_columnas + 1):
        letra = get_column_letter(numero_columna)

        if anchos and numero_columna in anchos:
            ws.column_dimensions[letra].width = anchos[numero_columna]
        else:
            ws.column_dimensions[letra].width = 15

    configurar_impresion(
        ws=ws,
        fila_encabezado=fila_encabezado,
        ultima_columna=ultima_columna,
        ultima_fila=fila_total,
    )

    return crear_respuesta_excel(
        workbook=wb,
        nombre_archivo=nombre_archivo,
    )