from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# =========================
# COLORES
# =========================

AZUL_TITULO = "1F4E78"
AZUL_CLARO = "DDEBF7"
VERDE_OK = "C6EFCE"
ROJO_ERROR = "FFC7CE"
AMARILLO_ALERTA = "FFEB9C"
GRIS_SUAVE = "D9E1F2"
BLANCO = "FFFFFF"
GRIS_BORDE = "B7B7B7"
GRIS_TEXTO = "666666"


# =========================
# RELLENOS
# =========================

RELLENO_TITULO = PatternFill(
    start_color=AZUL_TITULO,
    end_color=AZUL_TITULO,
    fill_type="solid",
)

RELLENO_OK = PatternFill(
    start_color=VERDE_OK,
    end_color=VERDE_OK,
    fill_type="solid",
)

RELLENO_ERROR = PatternFill(
    start_color=ROJO_ERROR,
    end_color=ROJO_ERROR,
    fill_type="solid",
)

RELLENO_ALERTA = PatternFill(
    start_color=AMARILLO_ALERTA,
    end_color=AMARILLO_ALERTA,
    fill_type="solid",
)

RELLENO_INFORMATIVO = PatternFill(
    start_color=AZUL_CLARO,
    end_color=AZUL_CLARO,
    fill_type="solid",
)

RELLENO_GRIS = PatternFill(
    start_color=GRIS_SUAVE,
    end_color=GRIS_SUAVE,
    fill_type="solid",
)


# =========================
# FUENTES
# =========================

FUENTE_TITULO = Font(
    bold=True,
    size=16,
    color=BLANCO,
)

FUENTE_ENCABEZADO = Font(
    bold=True,
    color=BLANCO,
)

FUENTE_NEGRITA = Font(
    bold=True,
)

FUENTE_GENERADO = Font(
    italic=True,
    color=GRIS_TEXTO,
)


# =========================
# ALINEACIONES
# =========================

ALINEACION_CENTRO = Alignment(
    horizontal="center",
    vertical="center",
)

ALINEACION_IZQUIERDA = Alignment(
    horizontal="left",
    vertical="center",
)

ALINEACION_DERECHA = Alignment(
    horizontal="right",
    vertical="center",
)

ALINEACION_ENVUELTA = Alignment(
    horizontal="left",
    vertical="center",
    wrap_text=True,
)


# =========================
# BORDES
# =========================

BORDE_FINO = Border(
    top=Side(style="thin", color=GRIS_BORDE),
    left=Side(style="thin", color=GRIS_BORDE),
    right=Side(style="thin", color=GRIS_BORDE),
    bottom=Side(style="thin", color=GRIS_BORDE),
)